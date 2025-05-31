#!/usr/bin/env python3
r"""
Multi-Strategy Python simulator for GER30EA with 4 variations, logs all trades,
and produces a single-sheet Excel with multiple sections:

(1)  ListAllTrades
(1B) Subtotals by (Strategy,Year)
(2)  TotalsByStrategy
(3)  TotalsByMonth
(4)  Year x Strategy Pivot
(5)  Major Metrics Table
(6)  Optimal Levels Analysis

We remove the “Total_hcXX” columns from the pivot (4). 
We keep “dist70_99”, “dist100_129”, “dist130_plus” for finalDist≥70,≥100,≥130. 
We also do a partial close if 16 min pass since peak & peakProfit in [32..35] => forcibly +32 pips. 
"""

import csv
import argparse
import logging
from datetime import datetime, time
from zoneinfo import ZoneInfo
import pathlib
import re
from collections import defaultdict, Counter
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    HAVE_OPENPYXL = True
except ImportError:
    HAVE_OPENPYXL = False

# ---------------------------
# Data Aggregation
# ---------------------------
def aggregator_factory():
    return {
        "count":0,
        "wins":0,
        "loses":0,
        "pips_list":[],
        "hc_50_list":[],"hc_60_list":[],"hc_70_list":[],"hc_80_list":[],"hc_90_list":[],"hc_100_list":[],

        "sumDD":0.0,
        "maxDD":0.0,
        "sumProfitPeak":0.0,
        "maxProfitPeak":0.0
    }

data = defaultdict(aggregator_factory)

def major_metrics_factory():
    return {
        "session1_list":[],
        "session2_list":[],
        "zone1_list":[],
        "zone2_list":[],
        "zone3_list":[],
        "zone4_list":[],
        "dist45_69_list":[],
        "dist70_99_list":[],
        "dist100_129_list":[],
        "dist130_plus_list":[]
    }

majorMetrics = defaultdict(major_metrics_factory)

open_dist_counter = Counter()
close_reason_counter= Counter()

# ---------------------------
# Common Config
# ---------------------------
TOLERANCE    = 9.0
GSL          = 40.0
SWEEP_CLOSE  = 179.0
OVERNIGHT_LIMIT = 200.0
MIDDAY_LIMIT    = 150.0

TIME_CLOSE_45_69   = 16
TIME_CLOSE_70_PLUS = 31

RETRACTION_TABLE = [
    (15.0,29.9,  0,18.0),
    (30.0,35.9,  1,0.0),
    (36.0,45.9,  2,0.0),
    (46.0,9999.0,-1,0.0)
]

DIST_LEVELS = [45,70,100,130]

SESSION1_START = (8,0)
SESSION1_END   = (12,30)
SESSION2_START = (14,30)
SESSION2_END   = (17,16)

SESSION1_ZONES= [
    ((8,16),(9,5),(9,31), 1,45,False),
    ((9,30),(9,45),(10,6),2,45,False),
    ((10,15),(10,45),(12,31),3,70,False),
    ((10,45),(11,45),(12,31),4,45,False),
]
SESSION2_ZONES= [
    ((14,46),(15,6),(17,16),1,45,False),
    ((15,15),(15,45),(17,16),2,70,False),
    ((15,45),(16,48),(17,16),3,45,False),
]

STRATEGY_NAMES = {
    1: "OriginalAll",
    2: "No15BE_EarliestZone",
    3: "No15BE_Any45Trade",
    4: "No15BE_AfterEarliestZone",
}

class SessionData:
    def __init__(self,name,stH,stM,enH,enM,zones):
        self.name=name
        self.startH=stH
        self.startM=stM
        self.endH=enH
        self.endM=enM
        self.active=False
        self.dayDate=None
        self.openPrice=None
        self.highPrice=None
        self.lowPrice=None
        self.allowed=True
        self.zones=zones

class Trade:
    def __init__(self, strategyID, direction, entryPrice, sessionID, zoneID,
                 forcedClose, openTime, finalDist, noCloseRules):
        self.strategyID   = strategyID
        self.direction    = direction
        self.entry        = entryPrice
        self.sessionID    = sessionID
        self.zoneID       = zoneID
        self.forcedClose  = forcedClose
        self.openTime     = openTime
        self.finalDist    = finalDist
        self.noCloseRules = noCloseRules

        self.peakHigh= entryPrice
        self.peakLow = entryPrice
        self.peakTime= openTime
        self.active= True

STRAT_STATE = {}

def init_strategy_states():
    for sid in (1,2,3,4):
        s1= SessionData("Session1", *SESSION1_START, *SESSION1_END, SESSION1_ZONES)
        s2= SessionData("Session2", *SESSION2_START, *SESSION2_END, SESSION2_ZONES)
        STRAT_STATE[sid]= {
            "session1": s1,
            "session2": s2,
            "activeTrades": {1: None, 2: None},
            "closedTrades": [],
            "overnightRef": None,
            "middayRef": None,
            "currentDate": None,
        }

# -------------------------------------------------------------------------
# CLI
# -------------------------------------------------------------------------
def parse_args():
    ap= argparse.ArgumentParser()
    ap.add_argument("--csv", required=True)
    ap.add_argument("--year_range", default="2024")
    ap.add_argument("--verbose", action="store_true")
    ap.add_argument("--src_tz", default="America/Chicago")
    ap.add_argument("--dst_tz", default="Europe/London")
    ap.add_argument("--out_dir", default=".")
    ap.add_argument("--excel", action="store_true")
    return ap.parse_args()

def setup_logging(verbose, out_dir, yr_tag=""):
    outp= pathlib.Path(out_dir)
    outp.mkdir(exist_ok=True)
    log_name= outp/ f"mq4_multi_{yr_tag}.log"
    logging.basicConfig(
        level=(logging.DEBUG if verbose else logging.INFO),
        format="%(asctime)s %(levelname)s: %(message)s",
        handlers=[
            logging.StreamHandler(),
            logging.FileHandler(log_name, mode="w", encoding="utf-8")
        ]
    )
    logging.info(f"Logging to {log_name}")

def parse_year_range(rng):
    if "-" in rng:
        s,e= rng.split("-")
        return range(int(s), int(e)+1)
    else:
        y= int(rng)
        return range(y,y+1)

# -------------------------------------------------------------------------
def minutes_diff(d1,d2):
    return int((d2-d1).total_seconds()//60)

def daily_reset_if_new_date(dt, sid):
    st= STRAT_STATE[sid]
    old= st["currentDate"]
    new_= dt.date()
    if old is None or old!=new_:
        st["currentDate"]= new_
        st["session1"].allowed= True
        st["session1"].active= False
        st["session2"].allowed= True
        st["session2"].active= False
        st["session1"].dayDate= new_
        st["session2"].dayDate= new_
        st["activeTrades"]={1: None, 2: None}
        logging.debug(f"[Strategy {sid}] daily reset for {new_}")

def in_time_window(dt,h1,m1, h2,m2):
    t= dt.time()
    st= time(h1,m1)
    ed= time(h2,m2)
    if ed<st:
        return (t>=st) or (t<ed)
    return (t>=st) and (t<ed)

def handle_session(dt, price, sid, sessNum):
    st= STRAT_STATE[sid]
    sess= st["session1"] if sessNum==1 else st["session2"]
    if sess.dayDate!= dt.date() and sess.active:
        end_session(sid, sessNum)
    if not sess.allowed:
        if sess.active:
            end_session(sid, sessNum)
        return

    if in_time_window(dt, sess.startH, sess.startM, sess.endH, sess.endM):
        if not sess.active:
            start_session(sid, sessNum, price, dt)
        else:
            if price> sess.highPrice: sess.highPrice= price
            if price< sess.lowPrice : sess.lowPrice= price
    else:
        if sess.active:
            end_session(sid, sessNum)

def start_session(sid, sessNum, price, dt):
    st= STRAT_STATE[sid]
    s_= st["session1"] if sessNum==1 else st["session2"]
    s_.active= True
    s_.openPrice= price
    s_.highPrice= price
    s_.lowPrice=  price
    logging.debug(f"[Strat{sid}] Session{sessNum} start @ {price:.1f}, date={dt.date()}")

def end_session(sid, sessNum):
    st= STRAT_STATE[sid]
    s_= st["session1"] if sessNum==1 else st["session2"]
    logging.debug(f"[Strat{sid}] Session{sessNum} end.")
    s_.active= False
    s_.openPrice=None
    s_.highPrice=None
    s_.lowPrice=None

def get_active_zone(dt, sid, sessNum):
    st= STRAT_STATE[sid]
    s_= st["session1"] if sessNum==1 else st["session2"]
    if not s_.active: return None
    for z_ in s_.zones:
        (stH,stM)= z_[0]
        (enH,enM)= z_[1]
        (fcH,fcM)= z_[2]
        zID=  z_[3]
        bD=   z_[4]
        nC=   z_[5]
        if in_time_window(dt, stH,stM, enH,enM):
            forced= datetime(dt.year, dt.month, dt.day, fcH,fcM, tzinfo=dt.tzinfo)
            return (forced, zID, bD, nC)
    return None

def calc_retraction(dt, price, sid, sessNum):
    st= STRAT_STATE[sid]
    s_= st["session1"] if sessNum==1 else st["session2"]
    if not s_.active or s_.openPrice is None:
        return 0.0
    if price>= s_.openPrice:
        return s_.highPrice- price
    return price- s_.lowPrice

def retraction_lookup(r):
    for (mn,mx,skip,shift) in RETRACTION_TABLE:
        if mn<=r<=mx:
            return (skip,shift)
    return (0,0)

def pick_final_distance(baseDist, skip, shift):
    try:
        baseIdx= DIST_LEVELS.index(baseDist)
    except:
        return None
    fIdx= baseIdx+ skip
    if fIdx>3: fIdx=3
    baseVal= DIST_LEVELS[baseIdx]
    nextVal= DIST_LEVELS[fIdx]
    shifted= baseVal+shift
    chosen= max(shifted, nextVal)
    if chosen> 130+ TOLERANCE:
        return None
    final=None
    for i, lv in enumerate(DIST_LEVELS):
        if abs(chosen-lv)<0.5:
            final= lv
            break
        if i<3 and chosen> lv and chosen< DIST_LEVELS[i+1]:
            final= DIST_LEVELS[i+1]
            break
        if i==3 and chosen> lv:
            final= lv
            break
    return final

def try_open_trade(dt, price, sid, sessNum):
    st= STRAT_STATE[sid]
    if st["activeTrades"][sessNum] is not None:
        return
    zone_ = get_active_zone(dt, sid, sessNum)
    if not zone_: return
    forcedC, zID, bD, nC= zone_

    r= calc_retraction(dt, price, sid, sessNum)
    skip, shift= retraction_lookup(r)
    if skip==-1:
        if sessNum==1: st["session1"].allowed=False
        else: st["session2"].allowed=False
        logging.debug(f"[Strat{sid}] session{sessNum} blocked≥46 => no trade")
        return

    s_= st["session1"] if sessNum==1 else st["session2"]
    direction= "BUY" if price< s_.openPrice else "SELL"
    dist= abs(price- s_.openPrice)
    if dist< bD:
        return
    if dist> bD+ TOLERANCE:
        return
    finalDist= pick_final_distance(bD, skip, shift)
    if finalDist is None: return
    if dist< finalDist: return
    if dist> finalDist+ TOLERANCE: return

    open_dist_counter[finalDist]+=1

    tr= Trade(sid, direction, price, sessNum, zID, forcedC, dt, finalDist, nC)
    st["activeTrades"][sessNum]= tr
    logging.debug(f"[Strat{sid}] OPEN s{sessNum} z{zID} {direction} ent={price:.1f}, finalDist={finalDist}, forced={forcedC}, noClose={nC}")

def manage_trade(dt, price, sid, sessNum):
    st= STRAT_STATE[sid]
    tr= st["activeTrades"][sessNum]
    if not tr or not tr.active: return

    if dt>= tr.forcedClose:
        close_trade(tr, price, "ForcedClose", dt)
        return
    dd= abs(price- tr.entry)
    if dd>= GSL:
        close_trade(tr, price, "GSL", dt)
        return
    if tr.direction=="BUY":
        if price> tr.peakHigh:
            tr.peakHigh= price
            tr.peakTime= dt
        if price< tr.peakLow:
            tr.peakLow= price
            tr.peakTime= dt
    else:
        if price< tr.peakLow:
            tr.peakLow= price
            tr.peakTime= dt
        if price> tr.peakHigh:
            tr.peakHigh= price
            tr.peakTime= dt
    if tr.noCloseRules and tr.strategyID!=1:
        return
    if tr.finalDist>=45 and tr.finalDist<70:
        skipBE= False
        if tr.strategyID==2:
            if tr.sessionID==1 and tr.zoneID==1:
                skipBE=True
        elif tr.strategyID==3:
            skipBE=True
        elif tr.strategyID==4:
            if not (tr.sessionID==1 and tr.zoneID==1):
                skipBE=True
        if tr.direction=="BUY":
            peakProfit= tr.peakHigh- tr.entry
        else:
            peakProfit= tr.entry- tr.peakLow
        elap= minutes_diff(tr.peakTime, dt)
        if 32<=peakProfit<=35 and elap>=16:
            forcedP= tr.entry+32 if tr.direction=="BUY" else tr.entry-32
            close_trade(tr, forcedP, "PartialClose32", dt)
            return
        if not skipBE:
            if tr.direction=="BUY":
                mae= tr.entry- tr.peakLow
            else:
                mae= tr.peakHigh- tr.entry
            if mae>=15 and abs(price- tr.entry)<1.0:
                close_trade(tr, price, "BreakEven-15", dt)
                return
        if elap>= TIME_CLOSE_45_69:
            close_trade(tr, price, "TimeClose16", dt)
    else:
        elap= minutes_diff(tr.peakTime, dt)
        if elap>= TIME_CLOSE_70_PLUS:
            close_trade(tr, price, "TimeClose31", dt)

def hard_close_pips(direction, entry, final_pips, peakProfit, level):
    if peakProfit>= level:
        return float(level)
    return float(final_pips)

def close_trade(tr, price, reason, dt):
    st= STRAT_STATE[tr.strategyID]
    pl= (price- tr.entry) if tr.direction=="BUY" else (tr.entry- price)
    if tr.direction=="BUY":
        peakProfitPips= tr.peakHigh- tr.entry
        peakDrawdownPips= tr.entry- tr.peakLow
    else:
        peakProfitPips= tr.entry- tr.peakLow
        peakDrawdownPips= tr.peakHigh- tr.entry

    hc_50=hard_close_pips(tr.direction,tr.entry,pl,peakProfitPips,50)
    hc_60=hard_close_pips(tr.direction,tr.entry,pl,peakProfitPips,60)
    hc_70=hard_close_pips(tr.direction,tr.entry,pl,peakProfitPips,70)
    hc_80=hard_close_pips(tr.direction,tr.entry,pl,peakProfitPips,80)
    hc_90=hard_close_pips(tr.direction,tr.entry,pl,peakProfitPips,90)
    hc_100=hard_close_pips(tr.direction,tr.entry,pl,peakProfitPips,100)

    c_= {
      "strategy": tr.strategyID,
      "session":  tr.sessionID,
      "zone":     tr.zoneID,
      "dir":      tr.direction,
      "entry":    tr.entry,
      "exit":     price,
      "pips":     pl,
      "hc_50":    hc_50,
      "hc_60":    hc_60,
      "hc_70":    hc_70,
      "hc_80":    hc_80,
      "hc_90":    hc_90,
      "hc_100":   hc_100,
      "reason":   reason,
      "openTime": tr.openTime,
      "closeTime":dt,
      "finalDist":tr.finalDist,

      "peakProfitPips": peakProfitPips,
      "peakDrawdownPips": peakDrawdownPips
    }
    st["closedTrades"].append(c_)
    close_reason_counter[(tr.strategyID,reason)]+=1

    logging.info(f"[Close] Strat{tr.strategyID} s{tr.sessionID} z{tr.zoneID} {tr.direction} "
                 f"ent={tr.entry:.1f} exit={price:.1f} pips={pl:.1f} reason={reason} "
                 f"(peakProfit={peakProfitPips:.1f},peakDD={peakDrawdownPips:.1f})")

    tr.active=False
    st["activeTrades"][tr.sessionID]= None

    sid= tr.strategyID
    y= dt.year
    m= dt.month
    b= data[(sid,y,m)]
    b["count"]+=1
    if pl>0: b["wins"]+=1
    else: b["loses"]+=1
    b["pips_list"].append(pl)
    b["hc_50_list"].append(hc_50)
    b["hc_60_list"].append(hc_60)
    b["hc_70_list"].append(hc_70)
    b["hc_80_list"].append(hc_80)
    b["hc_90_list"].append(hc_90)
    b["hc_100_list"].append(hc_100)
    b["sumDD"] += peakDrawdownPips
    if peakDrawdownPips> b["maxDD"]:
        b["maxDD"]= peakDrawdownPips
    b["sumProfitPeak"]+= peakProfitPips
    if peakProfitPips> b["maxProfitPeak"]:
        b["maxProfitPeak"]= peakProfitPips

    # major metrics
    mm= majorMetrics[(sid,y)]
    if tr.sessionID==1:
        mm["session1_list"].append(pl)
    else:
        mm["session2_list"].append(pl)
    z_= tr.zoneID
    if z_>=1 and z_<=4:
        mm[f"zone{z_}_list"].append(pl)
    d_= tr.finalDist
    if 45<= d_<70:
        mm["dist45_69_list"].append(pl)
    elif 70<= d_<100:
        mm["dist70_99_list"].append(pl)
    elif 100<= d_<130:
        mm["dist100_129_list"].append(pl)
    else:
        if d_>=130:
            mm["dist130_plus_list"].append(pl)

def check_sweeps(dt, price, sid, sessNum):
    st= STRAT_STATE[sid]
    tr= st["activeTrades"][sessNum]
    if not tr or not tr.active:
        return
    s_= st["session1"] if sessNum==1 else st["session2"]
    if s_.active and s_.openPrice is not None:
        dist= abs(price- s_.openPrice)
        if dist>= SWEEP_CLOSE:
            close_trade(tr, price, f"Sweep≥{SWEEP_CLOSE}", dt)

def check_outside_sessions(dt, price, sid, sessNum):
    s1_in= in_time_window(dt, *SESSION1_START, *SESSION1_END)
    s2_in= in_time_window(dt, *SESSION2_START, *SESSION2_END)
    if not s1_in and not s2_in:
        st= STRAT_STATE[sid]
        tr= st["activeTrades"][sessNum]
        if tr and tr.active:
            close_trade(tr, price, "OutsideSessions", dt)

def check_volatility(dt, price, sid):
    st= STRAT_STATE[sid]
    if dt.hour==17 and dt.minute==16:
        st["overnightRef"]= price
    if dt.hour==8 and dt.minute==0:
        if st["overnightRef"] is not None:
            if abs(price- st["overnightRef"])>= OVERNIGHT_LIMIT:
                st["session1"].allowed= False
                logging.debug(f"[Strat {sid}] session1 blocked≥200 overnight")
        st["overnightRef"]= None
    if dt.hour==12 and dt.minute==0:
        st["middayRef"]= price
    if dt.hour==14 and dt.minute==30:
        if st["middayRef"] is not None:
            if abs(price- st["middayRef"])>= MIDDAY_LIMIT:
                st["session2"].allowed= False
                logging.debug(f"[Strat {sid}] session2 blocked≥150 midday")
        st["middayRef"]= None

def run_backtest(csv_file, yrs, src_tz, dst_tz, produce_excel=False, out_dir="."):
    init_strategy_states()
    from zoneinfo import ZoneInfo
    tzSrc= ZoneInfo(src_tz)
    tzDst= ZoneInfo(dst_tz)

    f= pathlib.Path(csv_file)
    if not f.is_file():
        logging.error(f"CSV file not found: {csv_file}")
        return

    with f.open("r", newline="") as fh:
        import csv
        rdr= csv.reader(fh, delimiter=';')
        naive= None
        for row in rdr:
            if len(row)<6:
                continue
            dstr= row[0].strip()
            tstr= row[1].strip()
            dtRaw= f"{dstr} {tstr}"
            try:
                naive= datetime.strptime(dtRaw, "%d/%m/%Y %H:%M:%S")
            except:
                continue
            dtSrc= naive.replace(tzinfo=tzSrc)
            dt= dtSrc.astimezone(tzDst)
            if dt.year not in yrs:
                continue
            try:
                op= float(row[2])
                hi= float(row[3])
                lo= float(row[4])
                cl= float(row[5])
            except:
                continue

            for sid in (1,2,3,4):
                daily_reset_if_new_date(dt, sid)

            for p in [lo, hi]:
                for sid in (1,2,3,4):
                    handle_session(dt, p, sid, 1)
                    handle_session(dt, p, sid, 2)
                    check_volatility(dt, p, sid)
                    check_sweeps(dt, p, sid, 1)
                    check_sweeps(dt, p, sid, 2)
                    check_outside_sessions(dt, p, sid, 1)
                    check_outside_sessions(dt, p, sid, 2)
                    try_open_trade(dt, p, sid, 1)
                    try_open_trade(dt, p, sid, 2)
                    manage_trade(dt, p, sid, 1)
                    manage_trade(dt, p, sid, 2)

        if naive:
            last_dt= naive.replace(tzinfo=tzSrc).astimezone(tzDst)
            for sid in (1,2,3,4):
                st= STRAT_STATE[sid]
                for sID,tr in st["activeTrades"].items():
                    if tr and tr.active:
                        close_trade(tr, tr.entry, "EndOfBacktest", last_dt)

    produce_summaries_and_excel_single_sheet(yrs, out_dir, produce_excel)

# -------------------------------------------------------------------------
def median(lst):
    if not lst:
        return 0.0
    arr= sorted(lst)
    n= len(arr)
    mid= n//2
    if n%2==1:
        return arr[mid]
    return (arr[mid-1]+ arr[mid])/2.0

def produce_summaries_and_excel_single_sheet(yrs, out_dir, produce_excel):
    if not produce_excel:
        logging.info("Excel not requested => skip.")
        return
    if not HAVE_OPENPYXL:
        logging.warning("openpyxl not installed => skip.")
        return
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    all_keys= list(data.keys())
    if not all_keys:
        logging.info("No trades => no excel.")
        return

    sids_in= sorted({k[0] for k in all_keys})
    yrs_in= sorted({k[1] for k in all_keys})
    multi_year_str= f"{min(yrs_in)}-{max(yrs_in)}"
    date_tag= datetime.now().strftime("%Y%m%d")

    wb= Workbook()
    ws= wb.active
    ws.title= "Summary"

    row=1

    # -------------------------------------------------------------------------
    # (1) ListAllTrades
    # -------------------------------------------------------------------------
    ws.cell(row=row,column=1,value="(1) ListAllTrades").font= Font(bold=True)
    row+=2

    heads_1= [
        "Strategy","Year","Month","Trades",
        "pips","hc_50","hc_60","hc_70","hc_80","hc_90","hc_100",
        "peakProfit","peakDrawdown","reason","openTime","closeTime"
    ]
    for c_i,h_ in enumerate(heads_1,1):
        ws.cell(row=row,column=c_i,value=h_).font= Font(bold=True)
    row+=1

    bigList= []
    for sid in sids_in:
        for cT in STRAT_STATE[sid]["closedTrades"]:
            Y_= cT["closeTime"].year
            M_= cT["closeTime"].month
            bigList.append((sid,Y_,M_, cT))
    bigList.sort(key=lambda x: (x[0], x[1], x[2], x[3]["closeTime"]))

    # aggregator for (1B)
    sub_1B= defaultdict(lambda: {
        "count":0,"pips":0.0,
        "hc_50":0.0,"hc_60":0.0,"hc_70":0.0,"hc_80":0.0,"hc_90":0.0,"hc_100":0.0
    })

    for (sid,yy,mm, cT) in bigList:
        c=1
        ws.cell(row=row,column=c,value=STRATEGY_NAMES[sid]); c+=1
        ws.cell(row=row,column=c,value=yy); c+=1
        ws.cell(row=row,column=c,value=mm); c+=1
        ws.cell(row=row,column=c,value=1); c+=1

        ws.cell(row=row,column=c,value=round(cT["pips"],1)); c+=1
        ws.cell(row=row,column=c,value=round(cT["hc_50"],1)); c+=1
        ws.cell(row=row,column=c,value=round(cT["hc_60"],1)); c+=1
        ws.cell(row=row,column=c,value=round(cT["hc_70"],1)); c+=1
        ws.cell(row=row,column=c,value=round(cT["hc_80"],1)); c+=1
        ws.cell(row=row,column=c,value=round(cT["hc_90"],1)); c+=1
        ws.cell(row=row,column=c,value=round(cT["hc_100"],1)); c+=1

        ws.cell(row=row,column=c,value=round(cT["peakProfitPips"],1)); c+=1
        ws.cell(row=row,column=c,value=round(cT["peakDrawdownPips"],1)); c+=1
        ws.cell(row=row,column=c,value=cT["reason"]); c+=1
        ws.cell(row=row,column=c,value=cT["openTime"].strftime("%Y-%m-%d %H:%M")); c+=1
        ws.cell(row=row,column=c,value=cT["closeTime"].strftime("%Y-%m-%d %H:%M")); c+=1

        sub_1B[(sid,yy)]["count"]+=1
        sub_1B[(sid,yy)]["pips"]+= cT["pips"]
        sub_1B[(sid,yy)]["hc_50"]+= cT["hc_50"]
        sub_1B[(sid,yy)]["hc_60"]+= cT["hc_60"]
        sub_1B[(sid,yy)]["hc_70"]+= cT["hc_70"]
        sub_1B[(sid,yy)]["hc_80"]+= cT["hc_80"]
        sub_1B[(sid,yy)]["hc_90"]+= cT["hc_90"]
        sub_1B[(sid,yy)]["hc_100"]+= cT["hc_100"]

        row+=1

    row+=2

    # -------------------------------------------------------------------------
    # (1B) Subtotals by (Strategy,Year)
    # -------------------------------------------------------------------------
    ws.cell(row=row,column=1,value="(1B) Subtotals by (Strategy, Year)").font=Font(bold=True)
    row+=2

    heads_1B= [
      "Strategy","Year","Trades","SumPips","SumHC50","SumHC60","SumHC70","SumHC80","SumHC90","SumHC100"
    ]
    for c_i, h_ in enumerate(heads_1B,1):
        ws.cell(row=row,column=c_i,value=h_).font= Font(bold=True)
    row+=1

    for (sid,yy), ag_ in sorted(sub_1B.items()):
        c=1
        ws.cell(row=row,column=c,value=STRATEGY_NAMES[sid]); c+=1
        ws.cell(row=row,column=c,value=yy); c+=1
        ws.cell(row=row,column=c,value=ag_["count"]); c+=1
        ws.cell(row=row,column=c,value=round(ag_["pips"],1)); c+=1
        ws.cell(row=row,column=c,value=round(ag_["hc_50"],1)); c+=1
        ws.cell(row=row,column=c,value=round(ag_["hc_60"],1)); c+=1
        ws.cell(row=row,column=c,value=round(ag_["hc_70"],1)); c+=1
        ws.cell(row=row,column=c,value=round(ag_["hc_80"],1)); c+=1
        ws.cell(row=row,column=c,value=round(ag_["hc_90"],1)); c+=1
        ws.cell(row=row,column=c,value=round(ag_["hc_100"],1)); c+=1
        row+=1

    row+=2

    # -------------------------------------------------------------------------
    # (2) Totals By Strategy
    # -------------------------------------------------------------------------
    ws.cell(row=row,column=1,value="(2) TotalsByStrategy").font=Font(bold=True)
    row+=2

    heads_2= [
        "Strategy","YearRange","Trades","NetPips",
        "Wins","Loses","WinRate",
        "AvgDrawdown","MaxDrawdown","AvgPeakProfit","MaxPeakProfit",
        "AvgPipsAll","MedPipsAll",
        "AvgWinsOnly","MedWinsOnly",
        "AvgLossOnly","MedLossOnly",
        "AvgHC_50","MedHC_50",
        "AvgHC_60","MedHC_60",
        "AvgHC_70","MedHC_70",
        "AvgHC_80","MedHC_80",
        "AvgHC_90","MedHC_90",
        "AvgHC_100","MedHC_100"
    ]
    for c_i,h_ in enumerate(heads_2,1):
        ws.cell(row=row,column=c_i,value=h_).font=Font(bold=True)
    row+=1

    def gather_multi_year_stats(sid):
        allp=[]
        hc50=[]
        hc60=[]
        hc70=[]
        hc80=[]
        hc90=[]
        hc100=[]
        tcount=0
        tw=0
        tl=0
        sDD=0.0
        mDD=0.0
        sPk=0.0
        mPk=0.0
        for (ss,yy,mm), b_ in data.items():
            if ss== sid:
                tcount+= b_["count"]
                tw    += b_["wins"]
                tl    += b_["loses"]
                sDD   += b_["sumDD"]
                if b_["maxDD"]> mDD: mDD= b_["maxDD"]
                sPk   += b_["sumProfitPeak"]
                if b_["maxProfitPeak"]> mPk: mPk= b_["maxProfitPeak"]

                allp.extend(b_["pips_list"])
                hc50.extend(b_["hc_50_list"])
                hc60.extend(b_["hc_60_list"])
                hc70.extend(b_["hc_70_list"])
                hc80.extend(b_["hc_80_list"])
                hc90.extend(b_["hc_90_list"])
                hc100.extend(b_["hc_100_list"])

        return {
          "count": tcount,
          "wins": tw,
          "loses": tl,
          "sumDD": sDD,
          "maxDD": mDD,
          "sumPk": sPk,
          "maxPk": mPk,
          "allp": allp,
          "hc_50": hc50, "hc_60": hc60, "hc_70": hc70,
          "hc_80": hc80, "hc_90": hc90, "hc_100": hc100
        }

    def do_avg_med(lst):
        if not lst: return (0.0,0.0)
        s_= sum(lst)
        avg_= s_/ len(lst)
        med_= median(lst)
        return (avg_, med_)

    for sid in sids_in:
        st_= gather_multi_year_stats(sid)
        c_= st_["count"]
        w_= st_["wins"]
        l_= st_["loses"]
        wr_= (w_/ c_*100) if c_>0 else 0
        avgDD_= st_["sumDD"]/ c_ if c_>0 else 0
        maxDD_= st_["maxDD"]
        avgPk_= st_["sumPk"]/ c_ if c_>0 else 0
        maxPk_= st_["maxPk"]

        allp= st_["allp"]
        sumAll= sum(allp)
        avgAll= sumAll/ len(allp) if allp else 0
        medAll= median(allp)
        wip= [x for x in allp if x>0]
        lop= [x for x in allp if x<=0]
        avgW= sum(wip)/ len(wip) if wip else 0
        medW= median(wip) if wip else 0
        avgL= sum(lop)/ len(lop) if lop else 0
        medL= median(lop) if lop else 0

        a50,m50= do_avg_med(st_["hc_50"])
        a60,m60= do_avg_med(st_["hc_60"])
        a70,m70= do_avg_med(st_["hc_70"])
        a80,m80= do_avg_med(st_["hc_80"])
        a90,m90= do_avg_med(st_["hc_90"])
        a100,m100=do_avg_med(st_["hc_100"])

        c=1
        ws.cell(row=row,column=c,value=STRATEGY_NAMES[sid]); c+=1
        ws.cell(row=row,column=c,value= multi_year_str); c+=1
        ws.cell(row=row,column=c,value= c_); c+=1
        ws.cell(row=row,column=c,value=round(sumAll,1)); c+=1
        ws.cell(row=row,column=c,value= w_); c+=1
        ws.cell(row=row,column=c,value= l_); c+=1
        ws.cell(row=row,column=c,value= round(wr_,1)); c+=1
        ws.cell(row=row,column=c,value= round(avgDD_,1)); c+=1
        ws.cell(row=row,column=c,value= round(maxDD_,1)); c+=1
        ws.cell(row=row,column=c,value= round(avgPk_,1)); c+=1
        ws.cell(row=row,column=c,value= round(maxPk_,1)); c+=1

        ws.cell(row=row,column=c,value= round(avgAll,1)); c+=1
        ws.cell(row=row,column=c,value= round(medAll,1)); c+=1
        ws.cell(row=row,column=c,value= round(avgW,1)); c+=1
        ws.cell(row=row,column=c,value= round(medW,1)); c+=1
        ws.cell(row=row,column=c,value= round(avgL,1)); c+=1
        ws.cell(row=row,column=c,value= round(medL,1)); c+=1

        for (av_, md_) in [(a50,m50),(a60,m60),(a70,m70),(a80,m80),(a90,m90),(a100,m100)]:
            ws.cell(row=row,column=c,value=round(av_,1)); c+=1
            ws.cell(row=row,column=c,value=round(md_,1)); c+=1

        row+=1

    row+=2

    # -------------------------------------------------------------------------
    # (3) TotalsByMonth
    # -------------------------------------------------------------------------
    ws.cell(row=row,column=1,value="(3) TotalsByMonth").font=Font(bold=True)
    row+=2

    heads_3= [
      "Strategy","YearRange","Month","Trades","NetPips","Wins","Loses","WinRate",
      "AvgDrawdown","MaxDrawdown","AvgPeakProfit","MaxPeakProfit",
      "AvgPipsAll","MedPipsAll","AvgWin","MedWin","AvgLoss","MedLoss",
      "AvgHC_50","MedHC_50",
      "AvgHC_60","MedHC_60",
      "AvgHC_70","MedHC_70",
      "AvgHC_80","MedHC_80",
      "AvgHC_90","MedHC_90",
      "AvgHC_100","MedHC_100",
    ]
    for c_i,h_ in enumerate(heads_3,1):
        ws.cell(row=row,column=c_i,value=h_).font=Font(bold=True)
    row+=1

    monthlyAgg= defaultdict(aggregator_factory)
    for (sid_, y_, m_), b_ in data.items():
        ma= monthlyAgg[(sid_, m_)]
        ma["count"]+= b_["count"]
        ma["wins"]+= b_["wins"]
        ma["loses"]+= b_["loses"]
        ma["pips_list"].extend(b_["pips_list"])
        ma["hc_50_list"].extend(b_["hc_50_list"])
        ma["hc_60_list"].extend(b_["hc_60_list"])
        ma["hc_70_list"].extend(b_["hc_70_list"])
        ma["hc_80_list"].extend(b_["hc_80_list"])
        ma["hc_90_list"].extend(b_["hc_90_list"])
        ma["hc_100_list"].extend(b_["hc_100_list"])

        ma["sumDD"]+= b_["sumDD"]
        if b_["maxDD"]> ma["maxDD"]:
            ma["maxDD"]= b_["maxDD"]
        ma["sumProfitPeak"]+= b_["sumProfitPeak"]
        if b_["maxProfitPeak"]> ma["maxProfitPeak"]:
            ma["maxProfitPeak"]= b_["maxProfitPeak"]

    def do_avg_med(lst):
        if not lst: return (0.0,0.0)
        s_= sum(lst)
        a_= s_/ len(lst)
        m_= median(lst)
        return (a_, m_)

    for sid in sids_in:
        for m_ in range(1,13):
            rec= monthlyAgg.get((sid,m_), None)
            if not rec or rec["count"]<=0:
                continue
            c_= rec["count"]
            w_= rec["wins"]
            l_= rec["loses"]
            wr_= (w_/ c_* 100) if c_>0 else 0
            avgDD_= rec["sumDD"]/ c_ if c_>0 else 0
            mxDD_= rec["maxDD"]
            avgPk_= rec["sumProfitPeak"]/ c_ if c_>0 else 0
            mxPk_= rec["maxProfitPeak"]

            allp= rec["pips_list"]
            sumAll= sum(allp)
            avgAll= sumAll/ len(allp) if allp else 0
            medAll= median(allp)
            wip= [x for x in allp if x>0]
            lop= [x for x in allp if x<=0]
            avgW= sum(wip)/ len(wip) if wip else 0
            medW= median(wip) if wip else 0
            avgL= sum(lop)/ len(lop) if lop else 0
            medL= median(lop) if lop else 0

            a50,m50= do_avg_med(rec["hc_50_list"])
            a60,m60= do_avg_med(rec["hc_60_list"])
            a70,m70= do_avg_med(rec["hc_70_list"])
            a80,m80= do_avg_med(rec["hc_80_list"])
            a90,m90= do_avg_med(rec["hc_90_list"])
            a100,m100=do_avg_med(rec["hc_100_list"])

            c=1
            ws.cell(row=row,column=c,value=STRATEGY_NAMES[sid]); c+=1
            ws.cell(row=row,column=c,value=multi_year_str); c+=1
            ws.cell(row=row,column=c,value=m_); c+=1
            ws.cell(row=row,column=c,value=c_); c+=1
            ws.cell(row=row,column=c,value=round(sumAll,1)); c+=1
            ws.cell(row=row,column=c,value=w_); c+=1
            ws.cell(row=row,column=c,value=l_); c+=1
            ws.cell(row=row,column=c,value=round(wr_,1)); c+=1
            ws.cell(row=row,column=c,value=round(avgDD_,1)); c+=1
            ws.cell(row=row,column=c,value=round(mxDD_,1)); c+=1
            ws.cell(row=row,column=c,value=round(avgPk_,1)); c+=1
            ws.cell(row=row,column=c,value=round(mxPk_,1)); c+=1

            ws.cell(row=row,column=c,value=round(avgAll,1)); c+=1
            ws.cell(row=row,column=c,value=round(medAll,1)); c+=1
            ws.cell(row=row,column=c,value=round(avgW,1)); c+=1
            ws.cell(row=row,column=c,value=round(medW,1)); c+=1
            ws.cell(row=row,column=c,value=round(avgL,1)); c+=1
            ws.cell(row=row,column=c,value=round(medL,1)); c+=1

            for (av_,md_) in [(a50,m50),(a60,m60),(a70,m70),(a80,m80),(a90,m90),(a100,m100)]:
                ws.cell(row=row,column=c,value=round(av_,1)); c+=1
                ws.cell(row=row,column=c,value=round(md_,1)); c+=1

            row+=1

    row+=2

    # -------------------------------------------------------------------------
    # (4) Pivot Year x Strategy (only final, no hc)
    # -------------------------------------------------------------------------
    ws.cell(row=row,column=1,value="(4) Year x Strategy Pivot (final only)").font= Font(bold=True)
    row+=2
    pivot_final= defaultdict(float)
    for (sid_,y_,m_), b_ in data.items():
        pivot_final[(sid_,y_)] += sum(b_["pips_list"])

    colStart=1
    ws.cell(row=row,column=colStart,value="Year").font= Font(bold=True)
    c_= colStart+1
    for sid_ in (1,2,3,4):
        ws.cell(row=row,column=c_, value=f"{STRATEGY_NAMES[sid_]}_final").font=Font(bold=True)
        c_+=1
    ws.cell(row=row,column=c_,value="Total_final").font=Font(bold=True)
    row+=1

    allYears= sorted({k[1] for k in data.keys()})
    for y_ in allYears:
        c_= colStart
        ws.cell(row=row,column=c_,value=y_)
        c_+=1
        yearTot=0.0
        for sid_ in (1,2,3,4):
            v_ = pivot_final.get((sid_,y_),0.0)
            ws.cell(row=row,column=c_,value=round(v_,1))
            c_+=1
            yearTot+=v_
        ws.cell(row=row,column=c_,value=round(yearTot,1))
        row+=1

    row+=2

    # -------------------------------------------------------------------------
    # (5) Major Metrics
    # -------------------------------------------------------------------------
    ws.cell(row=row,column=1,value="(5) Major Metrics Table").font=Font(bold=True)
    row+=2

    rowHeaders= [
      "Session1","Session2",
      "Zone1","Zone2","Zone3","Zone4",
      "Dist45_69","Dist70_99","Dist100_129","Dist130_plus"
    ]
    ws.cell(row=row,column=1,value="Strategy").font=Font(bold=True)
    ws.cell(row=row,column=2,value="MetricName").font=Font(bold=True)
    c_=3
    sorted_yrs= sorted(yrs_in)
    for y_ in sorted_yrs:
        ws.cell(row=row,column=c_, value=str(y_)).font=Font(bold=True)
        c_+=1
    ws.cell(row=row,column=c_,value="All").font=Font(bold=True)
    row+=1

    def mm_sum(lst):
        return sum(lst) if lst else 0.0

    mmAllSid= defaultdict(major_metrics_factory)

    for sid in sids_in:
        for rH in rowHeaders:
            ws.cell(row=row,column=1,value=STRATEGY_NAMES[sid])
            ws.cell(row=row,column=2,value=rH)
            c_=3
            combined= []
            for y_ in sorted_yrs:
                mmX= majorMetrics.get((sid,y_), None)
                if not mmX:
                    val=0.0
                else:
                    key_= rH.lower()+"_list"
                    if key_ not in mmX:
                        val=0.0
                    else:
                        arr= mmX[key_]
                        val= mm_sum(arr)
                        combined.extend(arr)
                ws.cell(row=row,column=c_,value=round(val,1))
                c_+=1
            totVal= round(mm_sum(combined),1)
            ws.cell(row=row,column=c_,value=totVal)
            mmAllSid[sid][key_].extend(combined)
            row+=1
        row+=1

    for sid in sids_in:
        ws.cell(row=row,column=1,value=f"{STRATEGY_NAMES[sid]}-All(Metrics)").font=Font(bold=True)
        row+=1
        for rH in rowHeaders:
            key_= rH.lower()+"_list"
            arr= mmAllSid[sid][key_]
            val_= mm_sum(arr)
            ws.cell(row=row,column=2,value=rH).font=Font(bold=True)
            ws.cell(row=row,column=3,value=round(val_,1)).font=Font(bold=True)
            row+=1
        row+=2

    # -------------------------------------------------------------------------
    # (6) Optimal Levels
    # -------------------------------------------------------------------------
    row+=1
    ws.cell(row=row,column=1,value="(6) Optimal Levels Analysis").font=Font(bold=True)
    row+=2

    heads_6= [
        "Strategy","Trades",
        "Hit_50%","Hit_60%","Hit_70%","Hit_80%","Hit_90%","Hit_100%",
        "Net@50","Net@60","Net@70","Net@80","Net@90","Net@100",
        "Best_TP","Best_Net","Median_DD","P75_DD","Suggested_SL_Range"
    ]
    for c_i,h_ in enumerate(heads_6,1):
        ws.cell(row=row,column=c_i,value=h_).font=Font(bold=True)
    row+=1

    def pct(a,b):
        return round(a/b*100,1) if b>0 else 0.0

    from statistics import median as stat_median

    for sid in sids_in:
        trades= STRAT_STATE[sid]["closedTrades"]
        n= len(trades)
        if n==0:
            continue
        hit_ct= {L:0 for L in(50,60,70,80,90,100)}
        net_if= {L:0.0 for L in(50,60,70,80,90,100)}
        dd_list= []
        for cT in trades:
            pk= cT["peakProfitPips"]
            dd_list.append(cT["peakDrawdownPips"])
            # net_if[L] => sum of cT["hc_L"]
            # if pk≥ L => we say we "hit" that L
            for L in (50,60,70,80,90,100):
                net_if[L]+= cT[f"hc_{L}"]
                if pk>= L:
                    hit_ct[L]+=1
        dd_list.sort()
        med_dd= stat_median(dd_list) if dd_list else 0
        i75= int(0.75* len(dd_list))
        p75= dd_list[i75] if i75<len(dd_list) else med_dd
        hint_= f"{round(med_dd)}–{round(p75)}"

        best_L, best_val= max(net_if.items(), key=lambda kv: kv[1])

        c=1
        ws.cell(row=row,column=c,value=STRATEGY_NAMES[sid]); c+=1
        ws.cell(row=row,column=c,value=n); c+=1
        for L in (50,60,70,80,90,100):
            ws.cell(row=row,column=c,value=pct(hit_ct[L], n))
            c+=1
        for L in (50,60,70,80,90,100):
            cell= ws.cell(row=row,column=c,value=round(net_if[L],1))
            if L==best_L:
                cell.fill= PatternFill(start_color="C6EFCE",end_color="C6EFCE",fill_type="solid")
            c+=1
        ws.cell(row=row,column=c,value=best_L).fill= PatternFill(start_color="C6EFCE",end_color="C6EFCE",fill_type="solid")
        c+=1
        ws.cell(row=row,column=c,value=round(best_val,1)).fill= PatternFill(start_color="C6EFCE",end_color="C6EFCE",fill_type="solid")
        c+=1
        ws.cell(row=row,column=c,value=round(med_dd,1)); c+=1
        ws.cell(row=row,column=c,value=round(p75,1)); c+=1
        ws.cell(row=row,column=c,value=hint_); c+=1
        row+=1

    outp= pathlib.Path(out_dir)/ f"results_{date_tag}.xlsx"
    wb.save(outp)
    logging.info(f"Excel file saved => {outp}")

def main():
    args= parse_args()
    yrs= parse_year_range(args.year_range)
    yr_tag= re.sub(r'[^0-9\-]+','_', args.year_range)
    setup_logging(args.verbose, args.out_dir, yr_tag)
    run_backtest(
        csv_file=args.csv,
        yrs= yrs,
        src_tz= args.src_tz,
        dst_tz= args.dst_tz,
        produce_excel= args.excel,
        out_dir= args.out_dir
    )
    logging.info("All years complete.")

    print("\n=== Diagnostics ===")
    print("Opened trades by finalDist:", open_dist_counter)
    print("Close reasons by strategy:")
    for (theSid, theReason), c_ in sorted(close_reason_counter.items()):
        print(f"  Strat{theSid} – {theReason}: {c_}")

    data.clear()
    open_dist_counter.clear()
    close_reason_counter.clear()
    majorMetrics.clear()

if __name__=="__main__":
    main()
