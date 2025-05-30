#!/usr/bin/env python3
r"""
Multi-Strategy Python simulator for GER30EA with 4 variations, logs all trades,
and produces a single-sheet Excel with 3 sections (ListAllTrades, TotalsByStrategy, TotalsByMonth),
while applying DST conversions from --src_tz to --dst_tz.

Variations:
1) Original (Strategy #1)
2) Original except NO -15 break-even for 08:16–09:05 zone
3) Original except NO -15 break-even for *any* 45–69.9 trades
4) Original except -15 break-even is used *only* in the earliest zone (08:16–09:05).
   For all other 45–69.9 trades (zones or session2), skip break-even.

Usage Example:
  python mq4_backtest_v2.py ^
    --csv "C:\path\to\dax-1m.csv" ^
    --year_range 2000-2025 ^
    --src_tz America/Chicago ^
    --dst_tz Europe/London ^
    --excel ^
    --verbose
"""

import csv
import argparse
import logging
from datetime import datetime, time
import math
from zoneinfo import ZoneInfo
import pathlib
import re
from collections import defaultdict, Counter

# Attempt to import openpyxl (for Excel output). If missing, we'll skip it.
try:
    import openpyxl
    from openpyxl.styles import Font
    HAVE_OPENPYXL = True
except ImportError:
    HAVE_OPENPYXL = False

# -------------------------------------------------------------------------
# GLOBAL AGGREGATOR for final stats
# data[(strategyID, year, month)] = {"count":0, "pips":0.0, "wins":0, "loses":0}
# -------------------------------------------------------------------------
data = defaultdict(lambda: {"count":0,"pips":0.0,"wins":0,"loses":0})

# Diagnostic counters
open_dist_counter = Counter()
close_reason_counter = Counter()

# -------------------------------------------------------------------------
# COMMON CONFIG (shared by all 4 strategies)
# -------------------------------------------------------------------------
TOLERANCE       = 9.0
GSL             = 40.0
SWEEP_CLOSE     = 179.0
OVERNIGHT_LIMIT = 200.0
MIDDAY_LIMIT    = 150.0

TIME_CLOSE_45_69   = 16
TIME_CLOSE_70_PLUS = 31

RETRACTION_TABLE = [
    (15.0, 29.9,  0, 18.0),
    (30.0, 35.9,  1,  0.0),
    (36.0, 45.9,  2,  0.0),
    (46.0, 9999.0, -1, 0.0),
]

DIST_LEVELS = [45, 70, 100, 130]

SESSION1_START = (8, 0)
SESSION1_END   = (12, 30)

SESSION2_START = (14, 30)
SESSION2_END   = (17, 16)

SESSION1_ZONES = [
    ((8,16),(9,5),(9,31), 1, 45, False),   # now break-even possible
    ((9,30),(9,45),(10,6),2, 45, False),
    ((10,15),(10,45),(12,31),3, 70, False),
    ((10,45),(11,45),(12,31),4, 45, False),
]
SESSION2_ZONES = [
    ((14,46),(15,6),(17,16),1, 45, False),
    ((15,15),(15,45),(17,16),2, 70, False),
    ((15,45),(16,48),(17,16),3, 45, False),
]

# We define 4 strategies
STRATEGY_NAMES = {
    1: "OriginalAll",
    2: "No15BE_EarliestZone",
    3: "No15BE_Any45Trade",
    4: "No15BE_AfterEarliestZone",
}

# -------------------------------------------------------------------------
# State for each strategy
# -------------------------------------------------------------------------
class SessionData:
    def __init__(self, name, stH, stM, endH, endM, zones):
        self.name      = name
        self.startH    = stH
        self.startM    = stM
        self.endH      = endH
        self.endM      = endM
        self.active    = False
        self.dayDate   = None
        self.openPrice = None
        self.highPrice = None
        self.lowPrice  = None
        self.allowed   = True
        self.zones     = zones

class Trade:
    def __init__(self, strategyID, direction, entryPrice, sessionID, zoneID,
                 forcedClose, openTime, finalDist, noCloseRules):
        self.strategyID   = strategyID
        self.direction    = direction  # "BUY" or "SELL"
        self.entry        = entryPrice
        self.sessionID    = sessionID
        self.zoneID       = zoneID
        self.forcedClose  = forcedClose
        self.openTime     = openTime
        self.finalDist    = finalDist
        self.noCloseRules = noCloseRules
        self.peakHigh     = entryPrice
        self.peakLow      = entryPrice
        self.peakTime     = openTime
        self.active       = True

STRAT_STATE = {}

def init_strategy_states():
    """Create data for strategies #1..#4."""
    for sid in (1,2,3,4):
        s1 = SessionData("Session1", *SESSION1_START, *SESSION1_END, SESSION1_ZONES)
        s2 = SessionData("Session2", *SESSION2_START, *SESSION2_END, SESSION2_ZONES)
        STRAT_STATE[sid] = {
            "session1": s1,
            "session2": s2,
            "activeTrades": {1: None, 2: None},
            "closedTrades": [],
            "overnightRef": None,
            "middayRef": None,
            "currentDate": None,
        }

# -------------------------------------------------------------------------
# ARG PARSING & LOGGING
# -------------------------------------------------------------------------
def parse_args():
    ap = argparse.ArgumentParser()
    ap.add_argument("--csv", required=True,
                    help="CSV with D/M/Y;H:M:S;O;H;L;C (in --src_tz).")
    ap.add_argument("--year_range", default="2024",
                    help="e.g. 2024 or 2020-2025 (default=2024)")
    ap.add_argument("--verbose", action="store_true")
    ap.add_argument("--src_tz", default="America/Chicago",
                    help="Time zone of CSV data. We'll shift to --dst_tz for the logic.")
    ap.add_argument("--dst_tz", default="Europe/London",
                    help="Time zone for actual trading logic (UK time).")
    ap.add_argument("--out_dir", default=".",
                    help="Directory for logs & excel output (default=current dir).")
    ap.add_argument("--excel", action="store_true",
                    help="If set, produce an Excel file with monthly/yearly aggregates.")
    return ap.parse_args()

def setup_logging(verbose, out_dir, yr_tag=""):
    outp = pathlib.Path(out_dir)
    outp.mkdir(exist_ok=True)
    log_name = outp / f"mq4_multi_{yr_tag}.log"
    logging.basicConfig(
        level=(logging.DEBUG if verbose else logging.INFO),
        format="%(asctime)s %(levelname)s: %(message)s",
        handlers=[
            logging.StreamHandler(),
            logging.FileHandler(log_name, mode="w", encoding="utf-8")
        ]
    )
    logging.info(f"Logging to {log_name}")

def parse_year_range(rng):
    if "-" in rng:
        s,e = rng.split("-")
        return range(int(s), int(e)+1)
    else:
        y= int(rng)
        return range(y,y+1)

# -------------------------------------------------------------------------
# HELPER: minutes_diff
# -------------------------------------------------------------------------
def minutes_diff(d1, d2):
    """Return integer minutes from d1 to d2 (d2 - d1)."""
    return int((d2 - d1).total_seconds() // 60)

# -------------------------------------------------------------------------
# DAILY RESET, SESSIONS, ZONES
# -------------------------------------------------------------------------
def daily_reset_if_new_date(dt, sid):
    st = STRAT_STATE[sid]
    cd = st["currentDate"]
    day_date = dt.date()
    if cd is None or cd!= day_date:
        st["currentDate"] = day_date
        st["session1"].allowed = True
        st["session1"].active = False
        st["session2"].allowed = True
        st["session2"].active = False
        st["session1"].dayDate = day_date
        st["session2"].dayDate = day_date
        st["activeTrades"] = {1: None, 2: None}
        logging.debug(f"[Strategy {sid}] daily reset for {day_date}")

def in_time_window(dt, h1,m1, h2,m2):
    t= dt.time()
    st= time(h1,m1)
    ed= time(h2,m2)
    if ed<st:
        return (t>=st) or (t<ed)
    else:
        return (t>=st) and (t<ed)

def handle_session(dt, price, sid, sessNum):
    st= STRAT_STATE[sid]
    if sessNum==1:
        sess= st["session1"]
    else:
        sess= st["session2"]

    if sess.dayDate != dt.date() and sess.active:
        end_session(sid, sessNum)

    if not sess.allowed:
        if sess.active:
            end_session(sid, sessNum)
        return

    if in_time_window(dt, sess.startH, sess.startM, sess.endH, sess.endM):
        if not sess.active:
            start_session(sid, sessNum, price, dt)
        else:
            if price> sess.highPrice:
                sess.highPrice= price
            if price< sess.lowPrice:
                sess.lowPrice= price
    else:
        if sess.active:
            end_session(sid, sessNum)

def start_session(sid, sessNum, price, dt):
    st= STRAT_STATE[sid]
    sess= st["session1"] if sessNum==1 else st["session2"]
    sess.active= True
    sess.openPrice= price
    sess.highPrice= price
    sess.lowPrice= price
    logging.debug(f"[Strat{sid}] Session{sessNum} start @ {price:.1f}, date={dt.date()}")

def end_session(sid, sessNum):
    st= STRAT_STATE[sid]
    sess= st["session1"] if sessNum==1 else st["session2"]
    logging.debug(f"[Strat{sid}] Session{sessNum} end.")
    sess.active= False
    sess.openPrice= None
    sess.highPrice= None
    sess.lowPrice= None

def get_active_zone(dt, sid, sessNum):
    st= STRAT_STATE[sid]
    sess= st["session1"] if sessNum==1 else st["session2"]
    if not sess.active:
        return None
    for z in sess.zones:
        (stH,stM)= z[0]
        (enH,enM)= z[1]
        (fcH,fcM)= z[2]
        zoneID   = z[3]
        baseDist = z[4]
        noClose  = z[5]
        if in_time_window(dt, stH,stM, enH,enM):
            forcedC = datetime(dt.year, dt.month, dt.day, fcH, fcM, tzinfo=dt.tzinfo)
            return (forcedC, zoneID, baseDist, noClose)
    return None

# -------------------------------------------------------------------------
# RETRACTIONS & TOLERANCE
# -------------------------------------------------------------------------
def calc_retraction(dt, price, sid, sessNum):
    st= STRAT_STATE[sid]
    sess= st["session1"] if sessNum==1 else st["session2"]
    if not sess.active or sess.openPrice is None:
        return 0.0
    if price >= sess.openPrice:
        return sess.highPrice - price
    else:
        return price - sess.lowPrice

def retraction_lookup(r):
    for (mn,mx, skip, shift) in RETRACTION_TABLE:
        if mn<=r<=mx:
            return (skip, shift)
    return (0,0)

def pick_final_distance(baseDist, skip, shift):
    try:
        baseIdx= DIST_LEVELS.index(baseDist)
    except ValueError:
        return None
    finalIdx= baseIdx + skip
    if finalIdx>3:
        finalIdx=3
    baseVal= DIST_LEVELS[baseIdx]
    nextVal= DIST_LEVELS[finalIdx]
    shifted= baseVal+ shift
    chosen= max(shifted, nextVal)
    if chosen> (130+TOLERANCE):
        return None

    final= None
    for i, lv in enumerate(DIST_LEVELS):
        if abs(chosen- lv)<0.5:
            final= lv
            break
        if i<3 and chosen> lv and chosen< DIST_LEVELS[i+1]:
            final= DIST_LEVELS[i+1]
            break
        if i==3 and chosen> lv:
            final= lv
            break
    return final

# -------------------------------------------------------------------------
# OPEN & MANAGE TRADES
# -------------------------------------------------------------------------
def try_open_trade(dt, price, sid, sessNum):
    st= STRAT_STATE[sid]
    if st["activeTrades"][sessNum] is not None:
        return

    z= get_active_zone(dt, sid, sessNum)
    if not z: return
    forcedC, zoneID, baseD, noClose= z

    r= calc_retraction(dt, price, sid, sessNum)
    skip, shift= retraction_lookup(r)
    if skip==-1:
        if sessNum==1: st["session1"].allowed= False
        else:          st["session2"].allowed= False
        logging.debug(f"[Strat{sid}] Session{sessNum} blocked≥46 => no trade")
        return

    sess= st["session1"] if sessNum==1 else st["session2"]
    direction= "BUY" if price< sess.openPrice else "SELL"
    dist= abs(price- sess.openPrice)
    if dist< baseD: return
    if dist> (baseD+ TOLERANCE): return

    finalDist= pick_final_distance(baseD, skip, shift)
    if finalDist is None: return
    if dist< finalDist: return
    if dist> (finalDist+ TOLERANCE): return

    open_dist_counter[finalDist] += 1

    tr= Trade(sid, direction, price, sessNum, zoneID, forcedC, dt, finalDist, noClose)
    st["activeTrades"][sessNum]= tr
    logging.debug(
        f"[Strat{sid}] OPEN s{sessNum} z{zoneID} {direction} ent={price:.1f} finalDist={finalDist}, forced={forcedC.time()}, noClose={noClose}"
    )
    logging.debug(f"Opened @{finalDist}p sid={sid} sess={sessNum} zone={zoneID}")

def manage_trade(dt, price, sid, sessNum):
    st= STRAT_STATE[sid]
    tr= st["activeTrades"][sessNum]
    if not tr or not tr.active:
        return

    if dt>= tr.forcedClose:
        close_trade(tr, price, "ForcedClose", dt)
        return

    dd= abs(price- tr.entry)
    if dd>= GSL:
        close_trade(tr, price, "GSL", dt)
        return

    # update peaks
    if tr.direction=="BUY":
        if price> tr.peakHigh:
            tr.peakHigh= price
            tr.peakTime= dt
        if price< tr.peakLow:
            tr.peakLow= price
            tr.peakTime= dt
    else:
        if price< tr.peakLow:
            tr.peakLow= price
            tr.peakTime= dt
        if price> tr.peakHigh:
            tr.peakHigh= price
            tr.peakTime= dt

    if tr.noCloseRules and tr.strategyID != 1:
        return

    if tr.finalDist>=45 and tr.finalDist<70:
        skipBreakEven= False
        if tr.strategyID==2:
            # skip only earliest zone => (session1, zone1)
            if tr.sessionID==1 and tr.zoneID==1:
                skipBreakEven= True
        elif tr.strategyID==3:
            skipBreakEven= True
        elif tr.strategyID==4:
            if not (tr.sessionID==1 and tr.zoneID==1):
                skipBreakEven= True

        if not skipBreakEven:
            # NEW: check *adverse* excursion (-15) and recovery to ~0
            if tr.direction == "BUY":
                mae = tr.entry - tr.peakLow  # maximum adverse excursion
            else:
                mae = tr.peakHigh - tr.entry

            if mae >= 15 and abs(price - tr.entry) < 1.0:
                close_trade(tr, price, "BreakEven-15", dt)
                return

        holdMins= minutes_diff(tr.peakTime, dt)
        if holdMins>= TIME_CLOSE_45_69:
            close_trade(tr, price, "TimeClose16", dt)
    else:
        # finalDist≥70 => 31-min
        holdMins= minutes_diff(tr.peakTime, dt)
        if holdMins>= TIME_CLOSE_70_PLUS:
            close_trade(tr, price, "TimeClose31", dt)

def close_trade(tr: Trade, price, reason, dt):
    st= STRAT_STATE[tr.strategyID]
    pl= (price- tr.entry) if tr.direction=="BUY" else (tr.entry- price)
    st["closedTrades"].append({
        "strategy": tr.strategyID,
        "session":  tr.sessionID,
        "zone":     tr.zoneID,
        "dir":      tr.direction,
        "entry":    tr.entry,
        "exit":     price,
        "pips":     pl,
        "reason":   reason,
        "openTime": tr.openTime,
        "closeTime": dt,
        "finalDist": tr.finalDist,
    })
    close_reason_counter[(tr.strategyID, reason)] += 1

    logging.info(
        f"[Close] Strat{tr.strategyID} s{tr.sessionID} z{tr.zoneID} {tr.direction} "
        f"ent={tr.entry:.1f} exit={price:.1f} pips={pl:.1f} reason={reason}",
    )
    tr.active= False
    STRAT_STATE[tr.strategyID]["activeTrades"][tr.sessionID]= None

    # aggregator => record stats in 'data'
    sid= tr.strategyID
    y= dt.year
    m= dt.month
    data[(sid, y, m)]["count"] +=1
    data[(sid, y, m)]["pips"]  += pl
    if pl>0:
        data[(sid, y, m)]["wins"] +=1
    else:
        data[(sid, y, m)]["loses"]+=1

# -------------------------------------------------------------------------
# SWEEPS & VOLATILITY
# -------------------------------------------------------------------------
def check_sweeps(dt, price, sid, sessNum):
    st= STRAT_STATE[sid]
    tr= st["activeTrades"][sessNum]
    if not tr or not tr.active:
        return
    sess= st["session1"] if sessNum==1 else st["session2"]
    if sess.active and sess.openPrice is not None:
        dist= abs(price- sess.openPrice)
        if dist>= SWEEP_CLOSE:
            close_trade(tr, price, f"Sweep≥{SWEEP_CLOSE}", dt)

def check_outside_sessions(dt, price, sid, sessNum):
    s1_in= in_time_window(dt, *SESSION1_START, *SESSION1_END)
    s2_in= in_time_window(dt, *SESSION2_START, *SESSION2_END)
    if not s1_in and not s2_in:
        st= STRAT_STATE[sid]
        tr= st["activeTrades"][sessNum]
        if tr and tr.active:
            close_trade(tr, price, "OutsideSessions", dt)

def check_volatility(dt, price, sid):
    st= STRAT_STATE[sid]
    if dt.hour==17 and dt.minute==16:
        st["overnightRef"]= price
    if dt.hour==8 and dt.minute==0:
        if st["overnightRef"] is not None:
            if abs(price- st["overnightRef"])>= OVERNIGHT_LIMIT:
                st["session1"].allowed= False
                logging.debug(f"[Strat {sid}] session1 blocked≥200 overnight")
        st["overnightRef"]= None

    if dt.hour==12 and dt.minute==0:
        st["middayRef"]= price
    if dt.hour==14 and dt.minute==30:
        if st["middayRef"] is not None:
            if abs(price- st["middayRef"])>= MIDDAY_LIMIT:
                st["session2"].allowed= False
                logging.debug(f"[Strat {sid}] session2 blocked≥150 midday")
        st["middayRef"]= None

# -------------------------------------------------------------------------
# MAIN BACKTEST
# -------------------------------------------------------------------------
def run_backtest(csv_file, yrs, src_tz, dst_tz, produce_excel=False, out_dir="."):
    init_strategy_states()

    tzSrc= ZoneInfo(src_tz)
    tzDst= ZoneInfo(dst_tz)

    f= pathlib.Path(csv_file)
    if not f.is_file():
        logging.error(f"CSV file not found: {csv_file}")
        return

    with f.open("r", newline="") as fh:
        rdr= csv.reader(fh, delimiter=';')
        naive= None
        for row in rdr:
            if len(row)<6:  # skip invalid
                continue
            dstr= row[0].strip()
            tstr= row[1].strip()
            dtRaw= f"{dstr} {tstr}"
            try:
                naive= datetime.strptime(dtRaw, "%d/%m/%Y %H:%M:%S")
            except:
                continue

            dtSrc= naive.replace(tzinfo=tzSrc)
            dt= dtSrc.astimezone(tzDst)
            if dt.year not in yrs:
                continue

            try:
                op= float(row[2])
                hi= float(row[3])
                lo= float(row[4])
                cl= float(row[5])
            except:
                continue

            # daily reset for each strategy
            for sid in (1,2,3,4):
                daily_reset_if_new_date(dt, sid)

            # we feed lo & hi
            for p in [lo, hi]:
                for sid in (1,2,3,4):
                    handle_session(dt, p, sid, 1)
                    handle_session(dt, p, sid, 2)
                    check_volatility(dt, p, sid)
                    check_sweeps(dt, p, sid, 1)
                    check_sweeps(dt, p, sid, 2)
                    check_outside_sessions(dt, p, sid, 1)
                    check_outside_sessions(dt, p, sid, 2)
                    try_open_trade(dt, p, sid, 1)
                    try_open_trade(dt, p, sid, 2)
                    manage_trade(dt, p, sid, 1)
                    manage_trade(dt, p, sid, 2)

        # end-of-file => close any open trades
        if naive:
            last_dt= naive.replace(tzinfo=tzSrc).astimezone(tzDst)
            for sid in (1,2,3,4):
                st= STRAT_STATE[sid]
                for sID,tr in st["activeTrades"].items():
                    if tr and tr.active:
                        close_trade(tr, tr.entry, "EndOfBacktest", last_dt)

    produce_summaries_and_excel_single_sheet(yrs, out_dir, produce_excel)


# -------------------------------------------------------------------------
# SINGLE-SHEET EXCEL WITH 3 SECTIONS
# -------------------------------------------------------------------------
def produce_summaries_and_excel_single_sheet(yrs, out_dir, produce_excel):
    """
    Creates a single Excel sheet "Summary" with 3 sections:
      (1) List All Trades with Cume Pips per (sid, year)
      (2) Totals By Strategy (all years combined)
      (3) Totals By Month (all years combined)

    Each section ends with a bold totals row for each strategy.
    """
    if not produce_excel:
        logging.info("Excel not requested. Skipping summary sheets.")
        return
    if not HAVE_OPENPYXL:
        logging.warning("openpyxl not installed => no Excel output.")
        return

    import openpyxl
    from openpyxl.styles import Font

    # If we have no trades, nothing to do
    all_keys = list(data.keys())  # (sid, year, month)
    if not all_keys:
        logging.info("No trades => no Excel output.")
        return

    # Gather sets
    sids_in = sorted(set(k[0] for k in all_keys))
    yrs_in  = sorted(set(k[1] for k in all_keys))
    multi_year_str = f"{min(yrs_in)}-{max(yrs_in)}"
    date_tag = datetime.now().strftime("%Y%m%d")

    # Create workbook & primary sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    # -----------------------------------------------------------
    # SECTION 1: List All Trades
    # -----------------------------------------------------------
    row = 1
    ws.cell(row=row, column=1, value="(1) List All Trades")
    row += 1

    headers_1 = [
        "Strategy",
        "Year",
        "Month",
        "Trades",
        "NetPips",
        "Wins",
        "Loses",
        "WinRate",
        "CumePipsYear"
    ]
    for col_i, hdr in enumerate(headers_1, start=1):
        ws.cell(row=row, column=col_i, value=hdr).font = Font(bold=True)
    row += 1

    # We track (sid,year)-> cume pips, so we can fill the "CumePipsYear" column
    cume_year_dict = defaultdict(float)
    # Also track totals to produce a bold line for each strategy at the end
    totals_1 = defaultdict(lambda: {"count": 0, "pips": 0.0, "wins": 0, "loses": 0})

    all_keys.sort()  # sort by (sid,year,month)
    for (sid, yy, mm) in all_keys:
        st_ = data[(sid, yy, mm)]
        c   = st_["count"]
        p   = st_["pips"]
        w   = st_["wins"]
        l   = st_["loses"]
        wr  = (w / c * 100) if c > 0 else 0

        # accumulate
        cume_year_dict[(sid, yy)] += p
        totals_1[sid]["count"] += c
        totals_1[sid]["pips"]  += p
        totals_1[sid]["wins"]  += w
        totals_1[sid]["loses"] += l

        ws.cell(row=row, column=1, value=STRATEGY_NAMES[sid])
        ws.cell(row=row, column=2, value=yy)
        ws.cell(row=row, column=3, value=mm)
        ws.cell(row=row, column=4, value=c)
        ws.cell(row=row, column=5, value=round(p, 1))
        ws.cell(row=row, column=6, value=w)
        ws.cell(row=row, column=7, value=l)
        ws.cell(row=row, column=8, value=round(wr, 1))
        ws.cell(row=row, column=9, value=round(cume_year_dict[(sid, yy)], 1))
        row += 1

    # Bold totals row for each strategy in section 1
    for sid in sorted(totals_1.keys()):
        t = totals_1[sid]
        c_ = t["count"]
        p_ = t["pips"]
        w_ = t["wins"]
        l_ = t["loses"]
        wr_ = (w_ / c_ * 100) if c_ > 0 else 0
        ws.cell(row=row, column=1, value=f"{STRATEGY_NAMES[sid]} Totals").font = Font(bold=True)
        ws.cell(row=row, column=4, value=c_).font = Font(bold=True)
        ws.cell(row=row, column=5, value=round(p_, 1)).font = Font(bold=True)
        ws.cell(row=row, column=6, value=w_).font = Font(bold=True)
        ws.cell(row=row, column=7, value=l_).font = Font(bold=True)
        ws.cell(row=row, column=8, value=round(wr_, 1)).font = Font(bold=True)
        row += 1

    row += 2  # blank line

    # -----------------------------------------------------------
    # SECTION 2: Totals By Strategy (multi-year)
    # -----------------------------------------------------------
    ws.cell(row=row, column=1, value="(2) Totals By Strategy")
    row += 1

    headers_2 = [
        "Strategy",
        "YearRange",
        "Trades",
        "NetPips",
        "Wins",
        "Loses",
        "WinRate",
        "CumePipsAllYears"
    ]
    for col_i, hdr in enumerate(headers_2, start=1):
        ws.cell(row=row, column=col_i, value=hdr).font = Font(bold=True)
    row += 1

    # We'll sum everything for each strategy across all years
    for sid in sorted(sids_in):
        sumCount = 0
        sumPips  = 0.0
        sumWins  = 0
        sumLoses = 0
        for (s2, y2, m2) in data.keys():
            if s2 == sid:
                st2 = data[(s2, y2, m2)]
                sumCount += st2["count"]
                sumPips  += st2["pips"]
                sumWins  += st2["wins"]
                sumLoses += st2["loses"]
        wr_ = (sumWins / sumCount * 100) if sumCount > 0 else 0

        ws.cell(row=row, column=1, value=STRATEGY_NAMES[sid])
        ws.cell(row=row, column=2, value=multi_year_str)
        ws.cell(row=row, column=3, value=sumCount)
        ws.cell(row=row, column=4, value=round(sumPips, 1))
        ws.cell(row=row, column=5, value=sumWins)
        ws.cell(row=row, column=6, value=sumLoses)
        ws.cell(row=row, column=7, value=round(wr_, 1))
        ws.cell(row=row, column=8, value=round(sumPips, 1))
        row += 1

    row += 2  # blank line

    # -----------------------------------------------------------
    # SECTION 3: Totals By Month (across all years)
    # -----------------------------------------------------------
    ws.cell(row=row, column=1, value="(3) Totals By Month")
    row += 1

    headers_3 = [
        "Strategy",
        "YearRange",
        "Month",
        "Trades",
        "NetPips",
        "Wins",
        "Loses",
        "WinRate",
        "CumePipsYear"
    ]
    for col_i, hdr in enumerate(headers_3, start=1):
        ws.cell(row=row, column=col_i, value=hdr).font = Font(bold=True)
    row += 1

    # We'll build a monthly aggregator: monthlyAgg[(sid,month)] = ...
    monthlyAgg = defaultdict(lambda: {"count":0, "pips":0.0, "wins":0, "loses":0})
    for (sid_, y_, m_) in data.keys():
        st_ = data[(sid_, y_, m_)]
        monthlyAgg[(sid_, m_)]["count"] += st_["count"]
        monthlyAgg[(sid_, m_)]["pips"]  += st_["pips"]
        monthlyAgg[(sid_, m_)]["wins"]  += st_["wins"]
        monthlyAgg[(sid_, m_)]["loses"] += st_["loses"]

    # We'll do a cume as we list months 1..12 for each strategy
    monthlyCume= defaultdict(float)

    for sid in sorted(sids_in):
        monthlyCume[sid] = 0.0
        for m_ in range(1, 13):
            st2= monthlyAgg.get((sid, m_), None)
            if not st2 or st2["count"]<=0:
                continue
            c_ = st2["count"]
            p_ = st2["pips"]
            w_ = st2["wins"]
            l_ = st2["loses"]
            wr_ = (w_/c_*100) if c_>0 else 0
            monthlyCume[sid]+= p_

            ws.cell(row=row, column=1, value=STRATEGY_NAMES[sid])
            ws.cell(row=row, column=2, value=multi_year_str)
            ws.cell(row=row, column=3, value=m_)
            ws.cell(row=row, column=4, value=c_)
            ws.cell(row=row, column=5, value=round(p_,1))
            ws.cell(row=row, column=6, value=w_)
            ws.cell(row=row, column=7, value=l_)
            ws.cell(row=row, column=8, value=round(wr_,1))
            ws.cell(row=row, column=9, value=round(monthlyCume[sid],1))
            row+=1

        # After listing that strategy's months, we add a bold totals row
        stratCount = 0
        stratPips  = 0.0
        stratWins  = 0
        stratLoses = 0
        for m_ in range(1,13):
            st3= monthlyAgg.get((sid, m_), None)
            if st3:
                stratCount += st3["count"]
                stratPips  += st3["pips"]
                stratWins  += st3["wins"]
                stratLoses += st3["loses"]

        wr_ = (stratWins/stratCount*100) if stratCount>0 else 0
        rowBold = row
        ws.cell(row=rowBold, column=1, value=f"{STRATEGY_NAMES[sid]} MonthTotals").font= Font(bold=True)
        ws.cell(row=rowBold, column=4, value=stratCount).font= Font(bold=True)
        ws.cell(row=rowBold, column=5, value=round(stratPips,1)).font= Font(bold=True)
        ws.cell(row=rowBold, column=6, value=stratWins).font= Font(bold=True)
        ws.cell(row=rowBold, column=7, value=stratLoses).font= Font(bold=True)
        ws.cell(row=rowBold, column=8, value=round(wr_,1)).font= Font(bold=True)
        ws.cell(row=rowBold, column=9, value=round(monthlyCume[sid],1)).font= Font(bold=True)
        row+=1

    # Finally, save to a time-stamped file
    outp = pathlib.Path(out_dir)/f"results_{date_tag}.xlsx"
    wb.save(outp)
    logging.info(f"Excel file saved => {outp}")

# -------------------------------------------------------------------------
def main():
    args = parse_args()
    yrs= parse_year_range(args.year_range)
    yr_tag= re.sub(r'[^0-9\-]+','_', args.year_range)
    setup_logging(args.verbose, args.out_dir, yr_tag)

    run_backtest(
        csv_file=args.csv,
        yrs=yrs,
        src_tz=args.src_tz,
        dst_tz=args.dst_tz,
        produce_excel=args.excel,
        out_dir=args.out_dir
    )
    logging.info("All years complete.")

    # -------------------------------------------------------------------------
    # MASTER SUMMARY (across all years)
    # -------------------------------------------------------------------------
    yr_range= parse_year_range(args.year_range)
    master_summary= defaultdict(lambda: {"pips":0, "count":0, "wins":0, "loses":0})
    monthly_summary= defaultdict(lambda: {"pips":0, "count":0, "wins":0, "loses":0})

    for yr in yr_range:
        print(f"Processing year {yr}...")
        yr_tag = str(yr)
        setup_logging(args.verbose, args.out_dir, yr_tag)
        run_backtest(
            csv_file=args.csv,
            yrs=range(yr, yr+1),
            src_tz=args.src_tz,
            dst_tz=args.dst_tz,
            produce_excel=args.excel,
            out_dir=args.out_dir
        )

        for sid in (1, 2, 3, 4):  # Update all loops to include Strategy #4
            closed = STRAT_STATE[sid]["closedTrades"]
            for trd in closed:
                key_year = (sid, trd["openTime"].year)
                key_month = (sid, trd["openTime"].month)

                pips = trd["pips"]
                master_summary[key_year]["pips"] += pips
                master_summary[key_year]["count"] += 1
                if pips >= 0:
                    master_summary[key_year]["wins"] += 1
                else:
                    master_summary[key_year]["loses"] += 1

                monthly_summary[key_month]["pips"] += pips
                monthly_summary[key_month]["count"] += 1
                if pips >= 0:
                    monthly_summary[key_month]["wins"] += 1
                else:
                    monthly_summary[key_month]["loses"] += 1

    outp = pathlib.Path(args.out_dir) / "summary_master.xlsx"
    wb = openpyxl.Workbook()

    # Sheet 1: Yearly Totals
    ws1 = wb.active
    ws1.title = "Yearly Totals"
    ws1.append(["Strategy", "Year", "Trades", "NetPips", "Wins", "Loses", "WinRate", "TotalPipsYear"])
    cumulative_pips_by_strategy = defaultdict(float)

    for (sid, yr), val in sorted(master_summary.items()):
        wr = (val["wins"] / val["count"] * 100) if val["count"] > 0 else 0
        cumulative_pips_by_strategy[sid] += val["pips"]
        ws1.append([
            STRATEGY_NAMES[sid],
            yr,
            val["count"],
            round(val["pips"], 1),
            val["wins"],
            val["loses"],
            round(wr, 1),
            round(cumulative_pips_by_strategy[sid], 1)
        ])

    # Sheet 2: Monthly Totals Cumulative
    ws2 = wb.create_sheet("Monthly Cumulative")
    ws2.append(["Strategy", "Month", "Trades", "NetPips", "Wins", "Loses", "WinRate", "TotalPipsMonthlyCume"])
    monthly_cume = defaultdict(float)
    for sid in (1, 2, 3, 4):  # Update to include Strategy #4
        for m in range(1, 13):
            key = (sid, m)
            val = monthly_summary[key]
            if val["count"] == 0:
                continue
            wr = (val["wins"] / val["count"] * 100) if val["count"] > 0 else 0
            monthly_cume[sid] += val["pips"]
            ws2.append([
                STRATEGY_NAMES[sid],
                m,
                val["count"],
                round(val["pips"], 1),
                val["wins"],
                val["loses"],
                round(wr, 1),
                round(monthly_cume[sid], 1)
            ])

    # Sheet 3: Year vs Strategy Overview
    ws3 = wb.create_sheet("Year_vs_Strategy")
    years = sorted(set(y for (_, y) in master_summary))
    ws3.append(["Year"] + [STRATEGY_NAMES[sid] for sid in (1, 2, 3, 4)] + ["TotalPipsYear"])
    for yr in years:
        row = [yr]
        total = 0
        for sid in (1, 2, 3, 4):  # Update to include Strategy #4
            val = master_summary.get((sid, yr), {"pips": 0})
            row.append(round(val["pips"], 1))
            total += val["pips"]
        row.append(round(total, 1))
        ws3.append(row)

    # Final total row
    final_row = ["TotalPipsAllYears"]
    grand_total = 0
    for sid in (1, 2, 3, 4):  # Update to include Strategy #4
        strategy_total = sum(v["pips"] for (s, _), v in master_summary.items() if s == sid)
        final_row.append(round(strategy_total, 1))
        grand_total += strategy_total
    final_row.append(round(grand_total, 1))
    ws3.append(final_row)

    wb.save(outp)
    print(f"Final summary saved to {outp}")
    logging.info("All years complete.")


if __name__=="__main__":
    main()

print("\n=== Diagnostics ===")
print("Opened trades by finalDist:", open_dist_counter)
print("Close reasons by strategy:")
for k, v in close_reason_counter.items():
    print(f"  Strat{k[0]:d} – {k[1]}: {v}")

# Prevent double-counting
data.clear()
open_dist_counter.clear()
close_reason_counter.clear()
